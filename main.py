import requests
from openpyxl import load_workbook
import pandas as pd
import time
import xlrd
import threading
import json
from selenium import webdriver

mall_name_list = []

total_crawl = 0


class core_1(threading.Thread):
    def __init__(self,mall_name_list):
        threading.Thread.__init__(self) 

        self.workbook_name = '업체.xlsx'
        self.wb = load_workbook(self.workbook_name)
        self.page = self.wb.active
        self.mall_name_list = mall_name_list
        
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36")
        self.options.add_experimental_option('excludeSwitches',['enable-automation'])
        self.options.add_argument('--disable-blink-features=AutomationControlled')

        self.driver = webdriver.Chrome(executable_path = "./chromedrivers/chromedriver_1.exe", options=self.options)

    def run(self):

        for name, row in category_num.iterrows() : 
            for pagingindex in range(1,100) : 
                if pagingindex % 4 == 1 : 

                    params = {
                        'sort': 'rel',
                        'pagingIndex': str(pagingindex),
                        'pagingSize': '80',
                        'viewType': 'list',
                        'productSet': str(productset),
                        'catId': '50000807',
                        'brand': '',
                        'maker': '',
                        'spec': '',
                        'mall': '',
                        'deliveryFee': '',
                        'deliveryTypeValue': '',
                        'iq': '',
                        'eq': '',
                        'xq': '',
                        'frm': 'NVSHCHK',
                        'window': '',
                    }

                    response = requests.get('https://search.shopping.naver.com/api/search/category/' + str(row["카테고리번호"]), params=params)

                
                    if response.status_code == 200 :

                        itemlist = response.json()
                        for i in itemlist['shoppingResult']['products']:
                            try : 
                                malladdress = i["mallInfoCache"]["bizplBaseAddr"]
                                businessno = i["mallInfoCache"]["businessNo"]
                                malllink = i["mallPcUrl"]
                                shoppingmallname = i["mallInfoCache"]["name"]
                                print(shoppingmallname)
                                reviews = i["reviewCountSum"]
                                if "smartstore" in malllink and mallname in self.mall_name_list == False: 

                                    self.driver.get(malllink)    

                                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[0].click()
                                    time.sleep(1)
                                    try : 
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    except : 
                                        self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[1].click()
                                        time.sleep(1)
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    
                                    self.page.append(["",shoppingmallname, mallname,malladdress,businessno,representative,gogek,email, reviews,row["세분류"]])
                                    self.wb.save(filename=self.workbook_name)
                                    print("correct")
                                    mall_name_list.append(mallname)
                        
                            except : 
                                print("pass")
                            
                    else : 
                        print("too many requests error")

class core_2(threading.Thread):
    def __init__(self,mall_name_list):
        threading.Thread.__init__(self) 

        self.workbook_name = '업체.xlsx'
        self.wb = load_workbook(self.workbook_name)
        self.page = self.wb.active
        self.mall_name_list = mall_name_list
        
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36")
        self.options.add_experimental_option('excludeSwitches',['enable-automation'])
        self.options.add_argument('--disable-blink-features=AutomationControlled')

        self.driver = webdriver.Chrome(executable_path = "./chromedrivers/chromedriver_2.exe", options=self.options)

    def run(self):

        for name, row in category_num.iterrows() : 
            for pagingindex in range(1,100) : 
                if pagingindex % 4 == 2 : 

                    params = {
                        'sort': 'rel',
                        'pagingIndex': str(pagingindex),
                        'pagingSize': '80',
                        'viewType': 'list',
                        'productSet': str(productset),
                        'catId': '50000807',
                        'brand': '',
                        'maker': '',
                        'spec': '',
                        'mall': '',
                        'deliveryFee': '',
                        'deliveryTypeValue': '',
                        'iq': '',
                        'eq': '',
                        'xq': '',
                        'frm': 'NVSHCHK',
                        'window': '',
                    }

                    response = requests.get('https://search.shopping.naver.com/api/search/category/' + str(row["카테고리번호"]), params=params)

                
                    if response.status_code == 200 :

                        itemlist = response.json()
                        for i in itemlist['shoppingResult']['products']:
                            try : 
                                malladdress = i["mallInfoCache"]["bizplBaseAddr"]
                                businessno = i["mallInfoCache"]["businessNo"]
                                malllink = i["mallPcUrl"]
                                shoppingmallname = i["mallInfoCache"]["name"]
                                print(shoppingmallname)
                                reviews = i["reviewCountSum"]
                                if "smartstore" in malllink and mallname in self.mall_name_list == False: 

                                    self.driver.get(malllink)    

                                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[0].click()
                                    time.sleep(1)
                                    try : 
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    except : 
                                        self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[1].click()
                                        time.sleep(1)
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    
                                    self.page.append(["",shoppingmallname, mallname,malladdress,businessno,representative,gogek,email, reviews,row["세분류"]])
                                    self.wb.save(filename=self.workbook_name)
                                    print("correct")
                                    mall_name_list.append(mallname)
                        
                            except : 
                                print("pass")
                            
                    else : 
                        print("too many requests error")

class core_3(threading.Thread):
    def __init__(self,mall_name_list):
        threading.Thread.__init__(self) 

        self.workbook_name = '업체.xlsx'
        self.wb = load_workbook(self.workbook_name)
        self.page = self.wb.active
        self.mall_name_list = mall_name_list
        
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36")
        self.options.add_experimental_option('excludeSwitches',['enable-automation'])
        self.options.add_argument('--disable-blink-features=AutomationControlled')

        self.driver = webdriver.Chrome(executable_path = "./chromedrivers/chromedriver_3.exe", options=self.options)

    def run(self):

        for name, row in category_num.iterrows() : 
            for pagingindex in range(1,100) : 
                if pagingindex % 4 == 3 : 
                    params = {
                        'sort': 'rel',
                        'pagingIndex': str(pagingindex),
                        'pagingSize': '80',
                        'viewType': 'list',
                        'productSet': str(productset),
                        'catId': '50000807',
                        'brand': '',
                        'maker': '',
                        'spec': '',
                        'mall': '',
                        'deliveryFee': '',
                        'deliveryTypeValue': '',
                        'iq': '',
                        'eq': '',
                        'xq': '',
                        'frm': 'NVSHCHK',
                        'window': '',
                    }

                    response = requests.get('https://search.shopping.naver.com/api/search/category/' + str(row["카테고리번호"]), params=params)
                
                    if response.status_code == 200 :

                        itemlist = response.json()
                        for i in itemlist['shoppingResult']['products']:
                            try : 
                                malladdress = i["mallInfoCache"]["bizplBaseAddr"]
                                businessno = i["mallInfoCache"]["businessNo"]
                                malllink = i["mallPcUrl"]
                                shoppingmallname = i["mallInfoCache"]["name"]
                                print(shoppingmallname)
                                reviews = i["reviewCountSum"]
                                if "smartstore" in malllink and mallname in self.mall_name_list == False: 

                                    self.driver.get(malllink)    

                                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[0].click()
                                    time.sleep(1)
                                    try : 
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    except : 
                                        self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[1].click()
                                        time.sleep(1)
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    
                                    self.page.append(["",shoppingmallname, mallname,malladdress,businessno,representative,gogek,email, reviews,row["세분류"]])
                                    self.wb.save(filename=self.workbook_name)
                                    print("correct")
                                    mall_name_list.append(mallname)
                        
                            except : 
                                print("pass")
                            
                    else : 
                        print("too many requests error")

class core_4(threading.Thread):
    def __init__(self,mall_name_list):
        threading.Thread.__init__(self) 

        self.workbook_name = '업체.xlsx'
        self.wb = load_workbook(self.workbook_name)
        self.page = self.wb.active
        self.mall_name_list = mall_name_list
        
        self.options = webdriver.ChromeOptions()
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.54 Safari/537.36")
        self.options.add_experimental_option('excludeSwitches',['enable-automation'])
        self.options.add_argument('--disable-blink-features=AutomationControlled')

        self.driver = webdriver.Chrome(executable_path = "./chromedrivers/chromedriver_4.exe", options=self.options)

    def run(self):

        for name, row in category_num.iterrows() : 
            for pagingindex in range(1,100) : 
                if pagingindex % 4 == 0 : 
                    params = {
                        'sort': 'rel',
                        'pagingIndex': str(pagingindex),
                        'pagingSize': '80',
                        'viewType': 'list',
                        'productSet': str(productset),
                        'catId': '50000807',
                        'brand': '',
                        'maker': '',
                        'spec': '',
                        'mall': '',
                        'deliveryFee': '',
                        'deliveryTypeValue': '',
                        'iq': '',
                        'eq': '',
                        'xq': '',
                        'frm': 'NVSHCHK',
                        'window': '',
                    }

                    response = requests.get('https://search.shopping.naver.com/api/search/category/' + str(row["카테고리번호"]),  params=params)

                
                    if response.status_code == 200 :

                        itemlist = response.json()
                        for i in itemlist['shoppingResult']['products']:
                            try : 
                                malladdress = i["mallInfoCache"]["bizplBaseAddr"]
                                businessno = i["mallInfoCache"]["businessNo"]
                                malllink = i["mallPcUrl"]
                                shoppingmallname = i["mallInfoCache"]["name"]
                                print(shoppingmallname)
                                reviews = i["reviewCountSum"]
                                if "smartstore" in malllink and mallname in self.mall_name_list == False: 

                                    self.driver.get(malllink)    

                                    self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                                    self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[0].click()
                                    time.sleep(1)
                                    try : 
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    except : 
                                        self.driver.find_elements_by_class_name("_3oMcQ3LMwm")[1].click()
                                        time.sleep(1)
                                        texts = self.driver.find_elements_by_class_name("_10PxysFyMd")
                                        mallname = texts[0].text
                                        representative = texts[1].text
                                        gogek = texts[5].text
                                        email = texts[6].text
                                    
                                    self.page.append(["",shoppingmallname, mallname,malladdress,businessno,representative,gogek,email, reviews,row["세분류"]])
                                    self.wb.save(filename=self.workbook_name)
                                    print("correct")
                                    mall_name_list.append(mallname)
                        
                            except : 
                                print("pass")
                    else : 
                        print("too many requests error")

result_df = pd.DataFrame(columns=["업체명", "쇼핑몰명","주소", "사업자번호", "대표자", "고객센터", "이메일", "상품리뷰수", "카테고리"])
result_df.to_excel("업체.xlsx", encoding="utf-8-sig")


print('1:네이버페이')
print('2:쇼핑윈도')
menunum = input("옵션을 선택해주세요.")
print('1:네이버 랭킹순')
print('2:낮은 가격순')
print('3:높은 가격순')
print('4:등록일순')
print('5:리뷰 많은순')
print('6:리뷰 좋은순')
sortnum = input("정렬옵션 선택해주세요.")
category_num = pd.read_excel("20220125.xlsx")
if int(menunum) == 1 :
    productset = "checkout"
elif int(menunum) == 2 :
    productset = "window"

if int(sortnum) == 1 :
    sort = "rel"
elif int(sortnum) == 2 :
    sort = "price_asc"
elif int(sortnum) == 3 :
    sort = "price_desc"
elif int(sortnum) == 4 :
    sort = "date"
elif int(sortnum) == 5 :
    sort = "review"
elif int(sortnum) == 6 :
    sort = "review_rel"

one_core = core_1(mall_name_list)
two_core = core_2(mall_name_list)
three_core = core_3(mall_name_list)
four_core = core_4(mall_name_list)
one_core.start()
two_core.start()
three_core.start()
four_core.start()